#database.py

import openpyxl

#По id получает "all, name, surname, username, job, hobby" из бд
def get_info(id_person, status):
    wb = openpyxl.load_workbook(filename = 'persons_data.xlsx')
    sheet = wb['person']
    
    #Узнаем column_status (если 0 - то all)
    column_status = 0
    for i in range(1, sheet.max_column+1):
        status_db = sheet.cell(row=1, column=i).value
        if status == status_db:
            column_status = i
            break
    
    # Узнаем row_db по id_person (если 0 - то Error)
    row_db = 0
    for i in range(2, sheet.max_row+1):
        id_person_db = sheet.cell(row=i, column=1).value
        if id_person == id_person_db:
            row_db = i
            break
    
    if column_status != 0:
        info = sheet.cell(row=row_db, column=column_status).value
    else:
        info = {}
        for i in range(2, sheet.max_column+1):
            value = str(sheet.cell(row=row_db, column=i).value)
            status = sheet.cell(row=1, column=i).value
            info[status] = value
    return info

#Получает id пользователя, выводит статус (admin, noname, person, adminandperson)
def check_status(id_person):
    wb = openpyxl.load_workbook(filename = 'persons_data.xlsx')
    sheet = wb['admin']
    for i in range(1, sheet.max_column+1):
        id_person_db = sheet.cell(row=1, column=i).value
        if id_person_db == id_person:
            return 'admin'
        
    sheet = wb['person']
    for i in range(2, sheet.max_row+1):
        id_person_db = sheet.cell(row=i, column=1).value
        if id_person_db == id_person:
            return 'person'
    return 'noname'
    
#Добавляет в бд данные о пользователе по id (status = name, surname, username, hobby, job)
def add_info(id_person, status, value):
    wb = openpyxl.load_workbook(filename = 'persons_data.xlsx')
    sheet = wb['person']
    
    #Узнаем column_status (если 0 - то all)
    column_status = 0
    for i in range(1, sheet.max_column+1):
        status_db = sheet.cell(row=1, column=i).value
        if status == status_db:
            column_status = i
            break
    if column_status == 0:
        print('Error in add_info(id_person, status, value)')
        return False

    row_person = 0 #Если человека нет в бд, то 0
    for i in range(2, sheet.max_row+1):
        person_id_db = sheet.cell(row=i, column=1).value
        if person_id_db == id_person:
            row_person = i
            
    if row_person == 0:
        row_person = sheet.max_row+1
        sheet.cell(row=row_person, column=1).value = id_person
    
    sheet.cell(row=row_person, column=column_status).value = str(value)
    wb.save('persons_data.xlsx')
    
    