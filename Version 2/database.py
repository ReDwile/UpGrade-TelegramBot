import openpyxl
import os
import shutil


def v_file():
    files = os.listdir('data/')
    return len(files)


def save_new_info(data_array):
    old_name = 'data/persons_data_' + str(v_file()) + '.xlsx'
    new_name = 'data/persons_data_' + str(v_file() + 1) + '.xlsx'
    shutil.copyfile(old_name, new_name)

    data_old = saved_id()
    for i in data_old:
        delete_data(i)

    for gro in data_array:
        for p in gro:
            add_info(p['id'], 'name', p['name'])
            add_info(p['id'], 'surname', p['surname'])
            add_info(p['id'], 'username', p['username'])
            add_info(p['id'], 'hobby', p['hobby'])
            add_info(p['id'], 'job', p['job'])
    print('Новая таблица добавлена')


def get_info(id_person, status):
    '''По id получает "all, name, surname, username, job, hobby" из бд'''
    filename_name = 'data/persons_data_' + str(v_file()) + '.xlsx'
    wb = openpyxl.load_workbook(filename=filename_name)
    sheet = wb['person']

    # Узнаем column_status (если 0 - то all)
    column_status = 0
    for i in range(1, sheet.max_column + 1):
        status_db = sheet.cell(row=1, column=i).value
        if status == status_db:
            column_status = i
            break

    # Узнаем row_db по id_person (если 0 - то Error)
    row_db = 0
    for i in range(2, sheet.max_row + 1):
        id_person_db = sheet.cell(row=i, column=1).value
        if id_person == id_person_db:
            row_db = i
            break

    if column_status != 0:
        # Убрать после исправления ошибки
        try:
            info = sheet.cell(row=row_db, column=column_status).value
        except:
            print(f'Возникла ошибка в /database.py строчка 57. row_db: {row_db}, column_status: {column_status}, id_person: {id_person}')
            info = sheet.cell(row=row_db, column=column_status).value
    else:
        info = {}
        for i in range(2, sheet.max_column + 1):
            # Убрать после исправления
            try:
                value = str(sheet.cell(row=row_db, column=i).value)
            except:
                print(f'Возникла ошибка в /database.py строчка 66. row: {row_db}, column: {i}, id_person: {id_person}')
                value = str(sheet.cell(row=row_db, column=i).value)
            status = sheet.cell(row=1, column=i).value
            info[status] = value
    return info


def add_info(id_person, status, value):
    '''Добавляет в бд данные о пользователе по id (status = name, surname, username, hobby, job)'''
    filename_name = 'data/persons_data_' + str(v_file()) + '.xlsx'
    wb = openpyxl.load_workbook(filename=filename_name)
    sheet = wb['person']

    # Узнаем column_status (если 0 - то all)
    column_status = 0
    for i in range(1, sheet.max_column + 1):
        status_db = sheet.cell(row=1, column=i).value
        if status == status_db:
            column_status = i
            break
    if column_status == 0:
        print('Error in add_info(id_person, status, value)')
        return False

    none_values = -1  # Чтобы заполнял пропуски
    row_person = 0  # Если человека нет в бд, то 0
    for i in range(2, sheet.max_row + 1):
        person_id_db = sheet.cell(row=i, column=1).value
        if none_values == -1 and person_id_db == None:
            none_values = i

        if person_id_db == id_person:
            row_person = i

    if row_person == 0:
        if none_values == -1:
            row_person = sheet.max_row + 1
        else:
            row_person = none_values  # Добавляет инфу нового человека в пустую строчку

        sheet.cell(row=row_person, column=1).value = id_person

    sheet.cell(row=row_person, column=column_status).value = str(value)
    wb.save(filename_name)


def delete_data(id_person):
    '''По id пользователя удаляет все данные из бд'''
    filename_name = 'data/persons_data_' + str(v_file()) + '.xlsx'
    wb = openpyxl.load_workbook(filename=filename_name)
    sheet = wb['person']

    # Найдем строчку, в которой расположена информация о пользователе
    for i in range(2, sheet.max_row + 1):
        id_person_db = sheet.cell(row=i, column=1).value
        if id_person_db == id_person:
            row_person = i
            break

    # Удаляем данные
    for i in range(1, sheet.max_column + 1):
        sheet.cell(row=row_person, column=i).value = None
    wb.save(filename_name)


def saved_id():
    '''Выдает массив id зарегистрированных людей'''
    filename_name = 'data/persons_data_' + str(v_file()) + '.xlsx'
    wb = openpyxl.load_workbook(filename=filename_name)
    sheet = wb['person']

    # Все зареганные пользователи в массиве
    person_array = []
    for i in range(2, sheet.max_row + 1):
        id_person = sheet.cell(row=i, column=1).value
        if id_person != None:
            person_array.append(id_person)
    return person_array


# Парcит из бд словари с людьми
def user_dicts():
    '''Парсит из бд и выдает 2 словаря (первый - полностью зареганный людей, второй - неполностью)'''

    all_users = {}
    not_all_users = {}
    users = saved_id()
    for user_id in users:
        user = get_info(user_id, 'all')
        if not 'None' in user.values():
            # Если пользователь полностью зарегался
            all_users[user_id] = user

        else:
            # Если пользователь зареган частично
            not_all_users[user_id] = user

    return all_users, not_all_users

# Сохраняет отзыв в текстовый файл
def feedback(id_person, text):
    string = str(get_info(id_person, 'name')) + ' ' + str(get_info(id_person, 'surname')) + ': ' + text
    with open('feedback.txt', 'a') as f:
        f.write(string + '\n')

# Сохраняет логи
def logi(text):
    string = text
    with open('logi.txt', 'a') as f:
        f.write(string + '\n')